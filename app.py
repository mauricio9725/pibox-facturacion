import streamlit as st
import pandas as pd
import clickhouse_connect
import json
import os
import bcrypt
import logging
from datetime import date
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
)
logger = logging.getLogger("pibox")

# ─────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────
PURPLE_HEX   = "6B21A8"
LIGHT_BG_HEX = "F3E8FF"
_USERS_DIR  = os.getenv("USERS_DIR", os.path.dirname(os.path.abspath(__file__)))
USERS_FILE  = os.path.join(_USERS_DIR, "users.json")
ROLES       = ["admin", "operaciones", "financiero", "cliente"]

MENU_ICONS = {
    "Prefactura Cliente":   "📊  Prefactura Cliente",
    "Prefactura Interna":   "📋  Prefactura Interna",
    "Data":                 "🗃️  Data",
    "Error Tracker":        "🚨  Error Tracker",
    "Gestión de Usuarios":  "👥  Gestión de Usuarios",
    "Configuración":        "⚙️  Configuración",
}
MENU_BY_ROL = {
    "admin":       ["Prefactura Cliente", "Prefactura Interna", "Data", "Error Tracker", "Gestión de Usuarios", "Configuración"],
    "operaciones": ["Prefactura Cliente", "Prefactura Interna", "Data", "Configuración"],
    "financiero":  ["Prefactura Cliente", "Prefactura Interna", "Data", "Error Tracker"],
    "cliente":     ["Prefactura Cliente", "Configuración"],
}

ERROR_CONTROLS = {
    "CONTROL_300":                 ("CONTROL_300_valor",                "💰 Control 300"),
    "CONTROL_CERO":                ("CONTROL_CERO_valor",               "⭕ GMV Cero"),
    "CONTROL_VALOR_ALTO":          ("CONTROL_VALOR_ALTO_valor",         "📈 Valor Alto"),
    "CONTROL_VALOR_BAJO":          ("CONTROL_VALOR_BAJO_valor",         "📉 Valor Bajo"),
    "Control_balanceados_company": ("Control_balanceados_company_diff", "🏢 Balance Company"),
    "Control_balanceados_driver":  ("Control_balanceados_driver_diff",  "🏍️ Balance Driver"),
    "Control_perdidas":            ("Control_perdidas_diff",            "🔴 Pérdidas"),
}

# ─────────────────────────────────────────────
# ESTILOS EXCEL
# ─────────────────────────────────────────────
PURPLE_FILL = PatternFill(start_color=PURPLE_HEX,  end_color=PURPLE_HEX,  fill_type="solid")
LIGHT_FILL  = PatternFill(start_color=LIGHT_BG_HEX, end_color=LIGHT_BG_HEX, fill_type="solid")
WHITE_FILL  = PatternFill(start_color="FFFFFF",     end_color="FFFFFF",     fill_type="solid")
TOTAL_FILL  = PatternFill(start_color="EDE9FE",     end_color="EDE9FE",     fill_type="solid")

_THIN = Side(style="thin", color="E5E7EB")
THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_MED = Side(style="medium", color=PURPLE_HEX)
MED_BORDER  = Border(left=_MED, right=_MED, top=_MED, bottom=_MED)

HEADER_FONT = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
DATA_FONT   = Font(name="Calibri", size=10)
BOLD_FONT   = Font(name="Calibri", bold=True, size=10)
TOTAL_FONT  = Font(name="Calibri", bold=True, size=12)
GRAY_FONT   = Font(name="Calibri", color="808080", size=9)

DATE_COLS  = {"Fecha_VERDADERA", "Date_Time", "Scheduled_Time"}
MONEY_COLS = {
    "GMV", "Package_Declared_Value", "Contraentrega",
    "Ganancia_piloto", "Ganancia_Corporativo", "Ganancia_Total",
    "VAL_AMOUNT_BOOKING_DRIVER_PAYMENT", "VAL_AMOUNT_COMMISSION_COMPANY_PAYMENT",
    "Final_Service_Cost", "Valor_final_con_Ajuste",
    "additional_company_final_cost", "Final_cost",
    "Additional_final_cost", "Dispute_final_cost", "Total_final_cost",
}
RIGHT_ALIGN_COLS = MONEY_COLS | {
    "Distancia_Recorrida", "estimated_traveled_distance",
    "traveled_distance", "Traveled_Time",
}
LEFT_ALIGN_COLS = {
    "Nombre_Compania", "Usuario_Tienda", "Package_Reference_Numbers",
    "Direccion_Salida", "Direccion_Entrega", "Ciudad", "KAM",
    "Nombre_Driver", "Document_Type", "Company", "cost_center",
}

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo",  6: "Junio",  7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

# ─────────────────────────────────────────────
# CSS GLOBAL
# ─────────────────────────────────────────────
APP_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

/* ── Base ───────────────────────────────── */
html, body, [class*="css"] { font-family: 'Inter', sans-serif !important; }

.main .block-container {
    padding-top: 2rem;
    background: linear-gradient(135deg, #F8F7FF 0%, #F3E8FF 100%);
    min-height: 100vh;
}

/* ── Sidebar ────────────────────────────── */
[data-testid="stSidebar"] > div:first-child {
    background: linear-gradient(180deg, #6B21A8 0%, #4C1D95 100%) !important;
    border-right: none !important;
}
/* Forzar texto blanco en TODO el sidebar */
[data-testid="stSidebar"],
[data-testid="stSidebar"] *:not(button) {
    color: white !important;
}
[data-testid="stSidebar"] hr {
    border-color: rgba(255,255,255,0.25) !important;
    margin: 0.5rem 0 !important;
}
/* Items del menú */
[data-testid="stSidebar"] [data-testid="stRadio"] label,
[data-testid="stSidebar"] [data-testid="stRadio"] label > div,
[data-testid="stSidebar"] [data-testid="stRadio"] label p,
[data-testid="stSidebar"] [data-testid="stRadio"] label span {
    color: rgba(255,255,255,0.88) !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] label {
    padding: 0.55rem 1rem !important;
    border-radius: 8px !important;
    margin: 2px 4px !important;
    cursor: pointer !important;
    transition: background 0.2s !important;
    font-size: 0.9rem !important;
    font-weight: 500 !important;
    display: flex !important;
    align-items: center !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] label:hover {
    background: rgba(255,255,255,0.15) !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] label:has(input:checked) {
    background: rgba(255,255,255,0.22) !important;
    border-left: 3px solid white !important;
    font-weight: 600 !important;
    color: white !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] label:has(input:checked) * {
    color: white !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] input[type="radio"] { display: none !important; }
[data-testid="stSidebar"] button {
    border: 1.5px solid rgba(255,255,255,0.55) !important;
    color: rgba(255,255,255,0.9) !important;
    background: transparent !important;
    border-radius: 8px !important;
}
[data-testid="stSidebar"] button:hover {
    background: rgba(255,255,255,0.15) !important;
    border-color: white !important;
    color: white !important;
}

/* ── Moto loader ────────────────────────── */
.moto-loader {
    display: flex; align-items: center; gap: 14px;
    padding: 1.2rem 1.5rem; background: white;
    border-radius: 12px; border-left: 4px solid #6B21A8;
    box-shadow: 0 2px 12px rgba(107,33,168,0.12);
    font-size: 1rem; color: #6B21A8; font-weight: 500;
    margin: 1rem 0;
}
.moto-anim {
    font-size: 1.8rem;
    display: inline-block;
    animation: motoRide 0.7s ease-in-out infinite alternate;
}
@keyframes motoRide {
    from { transform: translateX(-6px) rotate(-3deg); }
    to   { transform: translateX(6px)  rotate(3deg);  }
}
.road {
    flex: 1; height: 3px;
    background: repeating-linear-gradient(90deg, #6B21A8 0, #6B21A8 16px, transparent 16px, transparent 28px);
    border-radius: 2px; animation: roadMove 0.4s linear infinite;
}
@keyframes roadMove { from { background-position: 0 0; } to { background-position: -28px 0; } }

/* ── Animaciones globales ───────────────── */
@keyframes fadeIn      { from { opacity:0; transform:translateY(16px); } to { opacity:1; transform:translateY(0); } }
@keyframes slideInLeft { from { opacity:0; transform:translateX(-20px); } to { opacity:1; transform:translateX(0); } }
@keyframes pulse       { 0%,100% { transform:scale(1); } 50% { transform:scale(1.04); } }
@keyframes float       { 0%,100% { transform:translateY(0); } 50% { transform:translateY(-8px); } }
@keyframes shimmer     { 0% { background-position:-400px 0; } 100% { background-position:400px 0; } }

.fade-in   { animation: fadeIn 0.5s ease both; }
.slide-in  { animation: slideInLeft 0.4s ease both; }

/* ── Fondo decorativo ───────────────────── */
.main .block-container::before {
    content: '';
    position: fixed; top: 0; right: 0;
    width: 420px; height: 420px;
    background: radial-gradient(circle, rgba(168,85,247,0.08) 0%, transparent 70%);
    pointer-events: none; z-index: 0;
    border-radius: 50%;
}

/* ── Login ──────────────────────────────── */
.login-wrapper {
    display: flex; flex-direction: column; align-items: center;
    justify-content: center; min-height: 72vh;
    animation: fadeIn 0.7s ease;
}
.login-illustration {
    margin-bottom: 1rem;
    animation: float 3s ease-in-out infinite;
}
.login-logo { font-size: 3.5rem; font-weight: 800; color: #6B21A8; letter-spacing: -1px; }
.login-sub  { font-size: 1.05rem; color: #6B7280; margin-top: -6px; margin-bottom: 1.5rem; }

/* ── Banner de módulo ───────────────────── */
.module-banner {
    background: linear-gradient(135deg, #6B21A8 0%, #9333EA 60%, #A855F7 100%);
    border-radius: 14px; padding: 1.2rem 1.8rem;
    display: flex; align-items: center; gap: 1rem;
    margin-bottom: 1.4rem; box-shadow: 0 4px 18px rgba(107,33,168,0.25);
    animation: fadeIn 0.4s ease;
}
.module-banner-icon { font-size: 2rem; }
.module-banner-text h2 { color: white !important; margin: 0 !important; font-size: 1.3rem !important; }
.module-banner-text p  { color: rgba(255,255,255,0.75); margin: 0; font-size: 0.85rem; }

/* ── Métricas ───────────────────────────── */
.metric-card {
    background: white;
    border-radius: 14px;
    padding: 1.3rem 1rem;
    box-shadow: 0 2px 14px rgba(107,33,168,0.09);
    border-top: 3px solid #6B21A8;
    text-align: center;
    transition: transform 0.25s, box-shadow 0.25s;
    height: 100%;
    animation: fadeIn 0.5s ease both;
}
.metric-card:nth-child(2) { animation-delay: 0.08s; }
.metric-card:nth-child(3) { animation-delay: 0.16s; }
.metric-card:nth-child(4) { animation-delay: 0.24s; }
.metric-card:hover { transform: translateY(-3px); box-shadow: 0 8px 24px rgba(107,33,168,0.18); }
.metric-icon   { font-size: 2rem; margin-bottom: 0.4rem; animation: float 4s ease-in-out infinite; }
.metric-value  { font-size: 1.55rem; font-weight: 700; color: #6B21A8; line-height: 1.1; }
.metric-label  { font-size: 0.75rem; color: #6B7280; margin-top: 4px; font-weight: 500; text-transform: uppercase; letter-spacing: 0.6px; }

/* ── Sección de filtros ─────────────────── */
.filter-header {
    display: flex; align-items: center; gap: 8px;
    font-size: 1rem; font-weight: 600; color: #6B21A8;
    margin-bottom: 0.8rem;
}

/* ── Vista previa ───────────────────────── */
.preview-resumen {
    background: white; border-radius: 10px;
    overflow: hidden; box-shadow: 0 2px 8px rgba(107,33,168,0.1);
    margin-bottom: 0.5rem;
}
.preview-resumen-header {
    background: #6B21A8; color: white;
    padding: 1rem 1.5rem; display: flex;
    justify-content: space-between; align-items: center;
}
.preview-company { font-weight: 700; font-size: 1.1rem; }
.preview-period  { font-size: 0.85rem; opacity: 0.85; }
.preview-resumen table { width: 100%; border-collapse: collapse; }
.preview-resumen td {
    padding: 0.7rem 1.5rem;
    border-bottom: 1px solid #F3E8FF;
    font-size: 0.95rem;
}
.preview-resumen tr:nth-child(odd) td  { background: #F3E8FF; }
.preview-resumen tr:nth-child(even) td { background: white; }
.preview-amount { text-align: right; font-weight: 600; color: #6B21A8; }
.preview-total  { font-weight: 700; font-size: 1rem; }
.preview-footer {
    text-align: center; color: #9CA3AF; font-size: 0.78rem;
    padding: 0.8rem 1.5rem; background: #FAFAFA;
}

/* ── Botón de descarga ──────────────────── */
.stDownloadButton > button {
    background: linear-gradient(135deg, #6B21A8, #9333EA) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    padding: 0.75rem 1.5rem !important;
    font-size: 1rem !important;
    font-weight: 600 !important;
    box-shadow: 0 4px 15px rgba(107,33,168,0.3) !important;
    transition: all 0.3s ease !important;
}
.stDownloadButton > button:hover {
    box-shadow: 0 6px 22px rgba(107,33,168,0.5) !important;
    transform: translateY(-2px) !important;
}

/* ── Botón primario ─────────────────────── */
button[kind="primary"], .stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #6B21A8, #9333EA) !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    transition: all 0.2s !important;
}
button[kind="primary"]:hover {
    box-shadow: 0 4px 14px rgba(107,33,168,0.4) !important;
    transform: translateY(-1px) !important;
}

/* ── Inputs focus ───────────────────────── */
input:focus, textarea:focus, select:focus,
[data-baseweb="input"] input:focus {
    border-color: #6B21A8 !important;
    box-shadow: 0 0 0 2px rgba(107,33,168,0.2) !important;
}

/* ── Titles ─────────────────────────────── */
h1 { color: #1E1B4B !important; font-weight: 700 !important; }
h2, h3 { color: #1E1B4B !important; font-weight: 600 !important; }

/* ── Tabs ───────────────────────────────── */
.stTabs [data-baseweb="tab-list"] { gap: 4px; }
.stTabs [data-baseweb="tab"] {
    border-radius: 8px 8px 0 0 !important;
    font-weight: 500 !important;
}
.stTabs [aria-selected="true"] {
    color: #6B21A8 !important;
    border-bottom-color: #6B21A8 !important;
}

/* ── Expanders ──────────────────────────── */
[data-testid="stExpander"] summary {
    font-weight: 600 !important;
    color: #1E1B4B !important;
}

/* ── Alerts ─────────────────────────────── */
.element-container .stAlert { border-radius: 10px !important; }

/* ── Footer ─────────────────────────────── */
.app-footer {
    text-align: center; padding: 2rem 0 1rem;
    color: #9CA3AF; font-size: 0.78rem; border-top: 1px solid #E5E7EB;
    margin-top: 3rem;
}

/* ── Ocultar branding Streamlit ──────────── */
#MainMenu { visibility: hidden; }
footer    { visibility: hidden; }
</style>
"""

# ─────────────────────────────────────────────
# GESTIÓN DE USUARIOS
# ─────────────────────────────────────────────
def _default_users() -> dict:
    admin_pw  = os.getenv("ADMIN_PASSWORD", "pibox2024")
    admin_usr = os.getenv("ADMIN_USER", "admin")
    force_pw  = os.getenv("ADMIN_PASSWORD") is not None
    usuarios  = [{
        "username": admin_usr,
        "password_hash": bcrypt.hashpw(admin_pw.encode(), bcrypt.gensalt()).decode(),
        "nombre_completo": "Administrador", "rol": "admin",
        "activo": True, "debe_cambiar_password": force_pw,
    }]
    # Usuarios adicionales definidos en st.secrets [extra_users]
    try:
        extra = st.secrets.get("extra_users", {})
        for uname, udata in extra.items():
            raw_pw = udata.get("password", "Pibox2024!")
            usuarios.append({
                "username": uname,
                "password_hash": bcrypt.hashpw(raw_pw.encode(), bcrypt.gensalt()).decode(),
                "nombre_completo": udata.get("nombre_completo", uname),
                "rol": udata.get("rol", "operaciones"),
                "activo": True,
                "debe_cambiar_password": True,
            })
    except Exception:
        pass
    return {"usuarios": usuarios}

def _gh_token() -> str:
    try:
        return st.secrets.get("GITHUB_TOKEN", "") or os.getenv("GITHUB_TOKEN", "")
    except Exception:
        return os.getenv("GITHUB_TOKEN", "")

def _gh_repo() -> str:
    try:
        return st.secrets.get("GITHUB_REPO", "") or os.getenv("GITHUB_REPO", "")
    except Exception:
        return os.getenv("GITHUB_REPO", "")

def _gh_load() -> dict:
    import requests as _req, base64 as _b64
    token, repo = _gh_token(), _gh_repo()
    if not token or not repo:
        return {}
    try:
        r = _req.get(
            f"https://api.github.com/repos/{repo}/contents/users.json",
            headers={"Authorization": f"token {token}", "Accept": "application/vnd.github.v3+json"},
            timeout=8,
        )
        if r.status_code == 200:
            return json.loads(_b64.b64decode(r.json()["content"]).decode())
    except Exception as e:
        logger.warning(f"GitHub load error: {e}")
    return {}

def _gh_save(data: dict) -> bool:
    import requests as _req, base64 as _b64
    token, repo = _gh_token(), _gh_repo()
    if not token or not repo:
        return False
    try:
        url = f"https://api.github.com/repos/{repo}/contents/users.json"
        headers = {"Authorization": f"token {token}", "Accept": "application/vnd.github.v3+json"}
        sha = ""
        r = _req.get(url, headers=headers, timeout=8)
        if r.status_code == 200:
            sha = r.json().get("sha", "")
        content = _b64.b64encode(json.dumps(data, indent=2, ensure_ascii=False).encode()).decode()
        payload = {"message": "update users", "content": content}
        if sha:
            payload["sha"] = sha
        r2 = _req.put(url, headers=headers, json=payload, timeout=8)
        return r2.status_code in (200, 201)
    except Exception as e:
        logger.warning(f"GitHub save error: {e}")
        return False

def _merge_extra_users(data: dict) -> dict:
    """Agrega usuarios de st.secrets[extra_users] si no existen ya."""
    try:
        extra = st.secrets.get("extra_users", {})
        if not extra:
            return data
        usernames = {u["username"] for u in data.get("usuarios", [])}
        changed = False
        for uname, udata in extra.items():
            if uname not in usernames:
                raw_pw = udata.get("password", "Pibox2024!")
                data["usuarios"].append({
                    "username": uname,
                    "password_hash": bcrypt.hashpw(raw_pw.encode(), bcrypt.gensalt()).decode(),
                    "nombre_completo": udata.get("nombre_completo", uname),
                    "rol": udata.get("rol", "operaciones"),
                    "activo": True,
                    "debe_cambiar_password": True,
                })
                changed = True
        if changed:
            _save_raw(data)
    except Exception:
        pass
    return data

def _load_raw() -> dict:
    gh = _gh_load()
    if gh:
        return _merge_extra_users(gh)
    if not os.path.exists(USERS_FILE):
        data = _default_users()
        _save_raw(data)
        logger.info("users.json creado con usuario admin por defecto")
        return data
    with open(USERS_FILE, "r", encoding="utf-8") as f:
        return _merge_extra_users(json.load(f))

def _save_raw(data: dict) -> None:
    _gh_save(data)
    try:
        with open(USERS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    except Exception:
        pass

def load_users() -> list:
    return _load_raw()["usuarios"]

def save_users(usuarios: list) -> None:
    _save_raw({"usuarios": usuarios})

def find_user(username: str):
    return next((u for u in load_users() if u["username"] == username), None)

def verify_password(username: str, password: str) -> bool:
    u = find_user(username)
    if not u or not u.get("activo", True):
        logger.warning(f"Login fallido: usuario='{username}' (no existe o inactivo)")
        return False
    ok = bcrypt.checkpw(password.encode(), u["password_hash"].encode())
    if ok:
        logger.info(f"Login exitoso: usuario='{username}' rol='{u.get('rol')}'")
    else:
        logger.warning(f"Login fallido: contraseña incorrecta para '{username}'")
    return ok

def change_password(username: str, new_password: str, clear_flag: bool = True) -> None:
    usuarios = load_users()
    for u in usuarios:
        if u["username"] == username:
            u["password_hash"] = bcrypt.hashpw(new_password.encode(), bcrypt.gensalt()).decode()
            if clear_flag: u["debe_cambiar_password"] = False
            break
    save_users(usuarios)

def create_user(username, nombre, password, rol):
    usuarios = load_users()
    if any(u["username"] == username for u in usuarios):
        return False, f"El usuario «{username}» ya existe."
    usuarios.append({
        "username": username,
        "password_hash": bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode(),
        "nombre_completo": nombre, "rol": rol,
        "activo": True, "debe_cambiar_password": True,
    })
    save_users(usuarios); return True, "Usuario creado exitosamente."

def update_user(username: str, updates: dict) -> None:
    usuarios = load_users()
    for u in usuarios:
        if u["username"] == username: u.update(updates); break
    save_users(usuarios)

def reset_password_admin(username: str, temp_password: str) -> None:
    change_password(username, temp_password, clear_flag=False)
    update_user(username, {"debe_cambiar_password": True})

def count_active_admins() -> int:
    return sum(1 for u in load_users() if u["rol"] == "admin" and u.get("activo", True))


# ─────────────────────────────────────────────
# BASE DE DATOS
# ─────────────────────────────────────────────
def _cfg(key: str, default: str = "") -> str:
    """Lee config desde st.secrets primero, luego env var."""
    try:
        v = st.secrets.get(key, None)
        if v:
            return str(v)
    except Exception:
        pass
    return os.getenv(key, default)

def _get_client():
    return clickhouse_connect.get_client(
        host=_cfg("CH_HOST"), port=int(_cfg("CH_PORT", "8443")),
        username=_cfg("CH_USER"), password=_cfg("CH_PASSWORD"),
        database=_cfg("CH_DATABASE"), secure=True, verify=False,
        compress=False, connect_timeout=30, send_receive_timeout=300,
    )

@st.cache_data(ttl=1800, show_spinner=False)
def load_companies() -> pd.DataFrame:
    try:
        return _get_client().query_df(
            "SELECT DISTINCT Nombre_Compania, Company_ID, NIT "
            "FROM picapmongoprod.reporte_facturacion "
            "ORDER BY Nombre_Compania"
        )
    except Exception as exc:
        st.error(f"❌ Error conectando a ClickHouse: {exc}")
        return pd.DataFrame(columns=["Nombre_Compania", "Company_ID", "NIT"])

def _sanitize(val: str) -> str:
    """Permite solo caracteres seguros para valores en queries."""
    safe_chars = set("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-_. @")
    return "".join(c for c in str(val) if c in safe_chars)

def query_data(search_mode: str, search_value, fi: date, ff: date):
    try:
        if search_mode == "nit":
            safe_nit = _sanitize(search_value)
            where = f"NIT = '{safe_nit}'"
        elif isinstance(search_value, list):
            ids = ", ".join(f"'{_sanitize(v)}'" for v in search_value)
            where = f"Company_ID IN ({ids})"
        else:
            where = f"Company_ID = '{_sanitize(search_value)}'"
        sql = (
            f"SELECT * FROM picapmongoprod.reporte_facturacion "
            f"WHERE {where} AND Fecha_VERDADERA BETWEEN '{fi}' AND '{ff}'"
        )
        logger.info(f"Query: {where} | {fi} → {ff}")
        df = _get_client().query_df(sql)
        # Quitar timezone de todas las columnas datetime para evitar error en Excel
        for col in df.select_dtypes(include=["datetimetz"]).columns:
            df[col] = df[col].dt.tz_localize(None)
        logger.info(f"Query OK: {len(df)} filas")
        return df, None
    except Exception as exc:
        logger.error(f"Query error: {exc}")
        return None, str(exc)

@st.cache_data(ttl=3600, show_spinner=False)
def query_error_tracker() -> pd.DataFrame:
    """Carga últimos 3 meses de pibox_error_tracker_2, solo registros con al menos un error."""
    try:
        conditions = " OR ".join(
            [f"{c} != 'Booking normal'" for c in ERROR_CONTROLS.keys()]
        )
        sql = (
            "SELECT * FROM picapmongoprod.pibox_error_tracker_2 "
            f"WHERE Fecha_VERDADERA >= today() - INTERVAL 3 MONTH "
            f"AND Company_ID != '64109ea8ff24f8002c7f267b' "
            f"AND ({conditions})"
        )
        df = _get_client().query_df(sql)
        for col in df.select_dtypes(include=["datetimetz"]).columns:
            df[col] = df[col].dt.tz_localize(None)
        logger.info(f"Error Tracker OK: {len(df)} filas")
        return df
    except Exception as exc:
        logger.error(f"Error Tracker error: {exc}")
        return pd.DataFrame()

def query_data_all(fi: date, ff: date):
    """Descarga TODA la data del período sin filtro de empresa."""
    try:
        sql = (
            f"SELECT * FROM picapmongoprod.reporte_facturacion "
            f"WHERE Fecha_VERDADERA BETWEEN '{fi}' AND '{ff}'"
        )
        logger.info(f"Query ALL: {fi} → {ff}")
        df = _get_client().query_df(sql)
        for col in df.select_dtypes(include=["datetimetz"]).columns:
            df[col] = df[col].dt.tz_localize(None)
        logger.info(f"Query ALL OK: {len(df)} filas")
        return df, None
    except Exception as exc:
        logger.error(f"Query ALL error: {exc}")
        return None, str(exc)


# ─────────────────────────────────────────────
# UTILIDADES EXCEL
# ─────────────────────────────────────────────
def _clean(val):
    if val is None: return ""
    try:
        if pd.isna(val): return ""
    except (TypeError, ValueError): pass
    if isinstance(val, pd.Timestamp):
        # Quitar timezone para compatibilidad con openpyxl
        if val.tzinfo is not None:
            val = val.tz_localize(None)
        return val.to_pydatetime()
    return val


def _show_moto_loader(msg: str):
    """Muestra animación de moto mientras carga. Llama .empty() cuando termine."""
    ph = st.empty()
    ph.markdown(
        f'<div class="moto-loader">'
        f'<span class="moto-anim">🏍️</span>'
        f'<div class="road"></div>'
        f'<span>{msg}</span>'
        f'</div>',
        unsafe_allow_html=True,
    )
    return ph

def _write_header(ws, headers: list, row: int = 1) -> None:
    ws.row_dimensions[row].height = 25
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = PURPLE_FILL; c.font = HEADER_FONT; c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")

def _write_data(ws, df: pd.DataFrame, start_row: int, col_map: dict = None) -> None:
    df_cols = list(col_map.values()) if col_map else list(df.columns)
    for r_idx, (_, row) in enumerate(df.iterrows()):
        fill = LIGHT_FILL if r_idx % 2 == 0 else WHITE_FILL
        ws.row_dimensions[start_row + r_idx].height = 18
        for c_idx, df_col in enumerate(df_cols, 1):
            raw = row[df_col] if df_col in df.columns else ""
            val = _clean(raw)
            cell = ws.cell(row=start_row + r_idx, column=c_idx, value=val)
            cell.fill = fill; cell.font = DATA_FONT; cell.border = THIN_BORDER
            if df_col in RIGHT_ALIGN_COLS:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif df_col in LEFT_ALIGN_COLS:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if df_col in DATE_COLS and val != "": cell.number_format = "DD/MM/YYYY"
            if df_col in MONEY_COLS: cell.number_format = '$#,##0'

def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len: int = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                try: max_len = max(max_len, len(str(cell.value)))
                except Exception: pass
        width: int = min(max(int(max_len) + 2, 10), 40)
        ws.column_dimensions[col_letter].width = width

def _freeze(ws) -> None:
    ws.freeze_panes = "A2"

def _finalize_sheet(ws) -> None:
    """Aplica freeze, auto_filter, sin cuadrícula."""
    _freeze(ws)
    _auto_width(ws)
    ws.sheet_view.showGridLines = False
    if ws.dimensions and ws.dimensions != "A1:A1":
        ws.auto_filter.ref = ws.dimensions


# ─────────────────────────────────────────────
# MAPAS DE COLUMNAS
# ─────────────────────────────────────────────
_PAQUETES_MAP = {
    "ID SERVICIO": "Booking_ID", "FECHA": "Fecha_VERDADERA",
    "EMPRESA": "Nombre_Compania", "USUARIO / TIENDA": "Usuario_Tienda",
    "# PAQUETE": "Package_Reference_Numbers", "VALOR DECLARADO": "Package_Declared_Value",
    "ESTADO PAQUETE": "Estado_Paquete", "CONTRAENTREGA": "Contraentrega",
    "DIRECCION": "Direccion_Entrega", "DISTANCIA DE VIAJE (m)": "Distancia_Recorrida",
    "TIEMPO DE VIAJE (s)": "Traveled_Time", "TIPO VEHICULO": "vt_name_es", "MONTO": "GMV",
}
_SERVICIOS_MAP = {
    "ID SERVICIO": "Booking_ID", "FECHA": "Fecha_VERDADERA",
    "EMPRESA": "Nombre_Compania", "TIPO VEHICULO": "vt_name_es", "MONTO": "GMV",
}
_PILOTOS_MAP = {
    "ID SERVICIO": "Booking_ID", "FECHA": "Fecha_VERDADERA",
    "EMPRESA": "Nombre_Compania", "TIPO VEHICULO": "vt_name_es",
    "ID Driver": "Driver_ID", "Nombre Driver": "Nombre_Driver",
    "Tipo Documento Driver": "Document_Type", "Número documento Driver": "COD_Identification",
    "Celular Driver": "Phone_Driver", "Email Driver": "Email_Driver",
    "Monto": "GMV", "Ganancia Corporativa": "Ganancia_Corporativo",
    "Ingreso Piloto": "VAL_AMOUNT_BOOKING_DRIVER_PAYMENT",
}
DATA_COLS_ORDER = [
    "Booking_ID", "Company_ID", "Fecha_VERDADERA", "Date_Time",
    "Scheduled_Time", "Ciudad", "Nombre_Compania", "Usuario_Tienda",
    "Package_Reference_Numbers", "Package_Declared_Value",
    "Estado_Paquete", "Contraentrega", "Direccion_Salida",
    "Direccion_Entrega", "Distancia_Recorrida",
    "estimated_traveled_distance", "traveled_distance",
    "Traveled_Time", "Final_Service_Cost", "Valor_final_con_Ajuste",
    "GMV", "Company", "Ganancia_piloto", "Ganancia_Corporativo",
    "Ganancia_Total", "KAM", "Driver_ID", "Nombre_Driver",
    "Document_Type", "COD_Identification",
    "VAL_AMOUNT_BOOKING_DRIVER_PAYMENT",
    "VAL_AMOUNT_COMMISSION_COMPANY_PAYMENT", "vt_name_es",
    "Payment_Type", "business_type", "additional_company_final_cost",
    "Final_cost", "Additional_final_cost", "Dispute_final_cost",
    "Total_final_cost", "express_service", "return_to_origin",
    "id_paquete", "NIT", "cost_center",
    "Phone_Driver", "Email_Driver",
]
DROP_SINDUP = {"Package_Reference_Numbers", "Package_Declared_Value", "Estado_Paquete", "Contraentrega", "id_paquete"}


# ─────────────────────────────────────────────
# CONSTRUCTORES DE HOJAS
# ─────────────────────────────────────────────
def _build_resumen(ws, df: pd.DataFrame, empresa: str, fi: date, ff: date, include_pilotos: bool = False) -> None:
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 45

    ws.merge_cells("A2:B2")
    c = ws["A2"]; c.value = "pibox"
    c.font = Font(name="Calibri", color=PURPLE_HEX, bold=True, size=22)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("C2:J2")
    c = ws["C2"]; c.value = "Informe facturación (costo del servicio de mensajería)"
    c.font = Font(name="Calibri", color=PURPLE_HEX, bold=True, size=14)
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[3].height = 8; ws.row_dimensions[4].height = 8

    svc_cols = [c for c in ["Booking_ID", "Fecha_VERDADERA", "Nombre_Compania", "vt_name_es", "GMV"] if c in df.columns]
    total_gmv = float(df[svc_cols].drop_duplicates()["GMV"].sum()) if "GMV" in svc_cols else 0.0
    periodo = f"{MESES_ES.get(fi.month, '')} {fi.year}"

    _write_header(ws, ["Campo", "Valor"], row=5)
    ws.column_dimensions["A"].width = 38; ws.column_dimensions["B"].width = 28

    tabla = [("Empresa", empresa), ("Periodo", periodo),
             ("Pago Servicios", total_gmv), ("Valor total de la factura:", total_gmv)]

    for r_idx, (label, val) in enumerate(tabla):
        fill = LIGHT_FILL if r_idx % 2 == 0 else WHITE_FILL
        ws.row_dimensions[6 + r_idx].height = 22
        lc = ws.cell(row=6 + r_idx, column=1, value=label)
        lc.border = THIN_BORDER
        lc.alignment = Alignment(horizontal="left", vertical="center")

        vc = ws.cell(row=6 + r_idx, column=2, value=val)
        vc.border = THIN_BORDER
        vc.alignment = Alignment(horizontal="right", vertical="center")

        if label == "Valor total de la factura:":
            lc.font = TOTAL_FONT; vc.font = TOTAL_FONT
            lc.fill = TOTAL_FILL; vc.fill = TOTAL_FILL
            lc.border = MED_BORDER; vc.border = MED_BORDER
        else:
            lc.font = DATA_FONT; lc.fill = fill
            vc.font = DATA_FONT; vc.fill = fill

        if label in ("Pago Servicios", "Valor total de la factura:"):
            vc.number_format = '$#,##0'

    # ── Sección análisis pilotos (solo Prefactura Interna) ───
    if not include_pilotos:
        footer_row = 17
        ws.merge_cells(f"A{footer_row}:J{footer_row}")
        ws.row_dimensions[footer_row].height = 28
        fc = ws[f"A{footer_row}"]
        fc.value = ("El plazo para dar el OK o para pedir ajustes es de 3 días "
                    "y de esta manera evitar demoras en el proceso de facturación")
        fc.font = GRAY_FONT
        fc.alignment = Alignment(horizontal="center", wrap_text=True)
        return

    _anal = [
        (11, "Pago a Pilotos",  "=SUM('Data Pilotos'!K:K)", '$#,##0.00', False),
        (12, "Utilidad",        "=B9-B11",                  '$#,##0',    True),
        (14, "Ganancia Corp",   "=SUM('Data Pilotos'!J:J)", '$#,##0.00', False),
        (15, "Dif",             "=B12-B14",                 '$#,##0',    True),
    ]
    for row_n, lbl, formula, num_fmt, is_total in _anal:
        ws.row_dimensions[row_n].height = 22
        lc = ws.cell(row=row_n, column=1, value=lbl)
        vc = ws.cell(row=row_n, column=2, value=formula)
        vc.number_format = num_fmt
        lc.border = THIN_BORDER; vc.border = THIN_BORDER
        lc.alignment = Alignment(horizontal="left",  vertical="center")
        vc.alignment = Alignment(horizontal="right", vertical="center")
        if is_total:
            lc.font = TOTAL_FONT; vc.font = TOTAL_FONT
            lc.fill = TOTAL_FILL; vc.fill = TOTAL_FILL
        else:
            lc.font = DATA_FONT;  vc.font = DATA_FONT
            lc.fill = LIGHT_FILL; vc.fill = LIGHT_FILL

    footer_row = 17
    ws.merge_cells(f"A{footer_row}:J{footer_row}")
    ws.row_dimensions[footer_row].height = 28
    fc = ws[f"A{footer_row}"]
    fc.value = ("El plazo para dar el OK o para pedir ajustes es de 3 días "
                "y de esta manera evitar demoras en el proceso de facturación")
    fc.font = GRAY_FONT
    fc.alignment = Alignment(horizontal="center", wrap_text=True)


def _build_paquetes(ws, df):
    _write_header(ws, list(_PAQUETES_MAP.keys()))
    _write_data(ws, df, 2, _PAQUETES_MAP)
    _finalize_sheet(ws)

def _build_servicios(ws, df):
    svc_df_cols = [c for c in _SERVICIOS_MAP.values() if c in df.columns]
    df_svc = df[svc_df_cols].drop_duplicates()
    _write_header(ws, list(_SERVICIOS_MAP.keys()))
    _write_data(ws, df_svc, 2, _SERVICIOS_MAP)
    _finalize_sheet(ws)

def _build_pilotos(ws, df):
    pil_cols = [c for c in _PILOTOS_MAP.values() if c in df.columns]
    df_pil = df[pil_cols].drop_duplicates()
    _write_header(ws, list(_PILOTOS_MAP.keys()))
    _write_data(ws, df_pil, 2, _PILOTOS_MAP)
    _finalize_sheet(ws)

def _build_data(ws, df):
    cols = [c for c in DATA_COLS_ORDER if c in df.columns]
    _write_header(ws, cols); _write_data(ws, df[cols], 2); _finalize_sheet(ws)

def _build_sin_duplicados(ws, df):
    cols = [c for c in DATA_COLS_ORDER if c in df.columns and c not in DROP_SINDUP]
    _write_header(ws, cols); _write_data(ws, df[cols].drop_duplicates(), 2); _finalize_sheet(ws)


# ─────────────────────────────────────────────
# GENERADORES DE EXCEL
# ─────────────────────────────────────────────
def _wb_to_bytes(wb) -> bytes:
    buf = BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

def gen_prefactura_cliente(df, empresa, fi, ff) -> bytes:
    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "Resumen"; _build_resumen(ws1, df, empresa, fi, ff)
    _build_paquetes(wb.create_sheet("Informe Paquetes"), df)
    _build_servicios(wb.create_sheet("Informe Servicios"), df)
    return _wb_to_bytes(wb)

def gen_prefactura_interna(df, empresa, fi, ff) -> bytes:
    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "Resumen"; _build_resumen(ws1, df, empresa, fi, ff, include_pilotos=True)
    _build_paquetes(wb.create_sheet("Informe Paquetes"), df)
    _build_servicios(wb.create_sheet("Informe Servicios"), df)
    _build_pilotos(wb.create_sheet("Data Pilotos"), df)
    return _wb_to_bytes(wb)

def gen_data_excel(df, empresa, fi, ff) -> bytes:
    cols = [c for c in DATA_COLS_ORDER if c in df.columns]
    cols_nodup = [c for c in cols if c not in DROP_SINDUP]
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df[cols].to_excel(writer, sheet_name="Data", index=False)
        df[cols_nodup].drop_duplicates().to_excel(writer, sheet_name="Sin Duplicados", index=False)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────
# FORMATEO DE MONEDA
# ─────────────────────────────────────────────
def _fmt_cop(value: float) -> str:
    return f"${value:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")


# ─────────────────────────────────────────────
# PÁGINAS
# ─────────────────────────────────────────────
def _page_login() -> None:
    import os as _os, base64 as _b64
    _base = _os.path.dirname(_os.path.abspath(__file__))
    _logo = next((_os.path.join(_base, f) for f in ["logo.png", "logo.jpeg", "logo.png.jpeg", "logo.jpg"] if _os.path.exists(_os.path.join(_base, f))), "")

    st.markdown("<div style='margin-top:8vh'></div>", unsafe_allow_html=True)
    _, col, _ = st.columns([1, 1, 1])
    with col:
        if _logo:
            with open(_logo, "rb") as _f:
                _enc = _b64.b64encode(_f.read()).decode()
            _ext = "jpeg" if _logo.endswith((".jpeg", ".jpg", ".png.jpeg")) else "png"
            st.markdown(
                f"<div style='text-align:center;margin-bottom:0.5rem'>"
                f"<img src='data:image/{_ext};base64,{_enc}' style='width:140px;max-width:100%;image-rendering:auto'/>"
                f"</div>",
                unsafe_allow_html=True,
            )
        st.markdown(
            "<p style='text-align:center;color:#6B7280;font-size:1rem;margin:0 0 1.4rem'>Portal de Prefacturación</p>",
            unsafe_allow_html=True,
        )
        with st.form("login_form"):
            username = st.text_input("Usuario", placeholder="Ingresa tu usuario")
            password = st.text_input("Contraseña", type="password", placeholder="Ingresa tu contraseña")
            submitted = st.form_submit_button("Iniciar sesión", use_container_width=True, type="primary")
        if submitted:
            if verify_password(username, password):
                u = find_user(username)
                st.session_state.update({
                    "authenticated": True, "username": username,
                    "user_display": u["nombre_completo"], "rol": u["rol"],
                    "debe_cambiar_password": u.get("debe_cambiar_password", False),
                })
                st.rerun()
            else:
                st.error("❌ Usuario o contraseña incorrectos, o cuenta inactiva.")


def _page_force_change_password() -> None:
    _, col, _ = st.columns([1, 1.4, 1])
    with col:
        st.markdown("<h2 style='color:#6B21A8'>🔐 Cambiar contraseña</h2>", unsafe_allow_html=True)
        st.info("⚠️ Por seguridad, debes establecer una nueva contraseña antes de continuar.")
        with st.form("force_pw"):
            nueva    = st.text_input("Nueva contraseña", type="password")
            confirmar = st.text_input("Confirmar nueva contraseña", type="password")
            submitted = st.form_submit_button("Establecer contraseña", use_container_width=True, type="primary")
        if submitted:
            if nueva != confirmar: st.error("Las contraseñas no coinciden.")
            elif len(nueva) < 6:   st.warning("Mínimo 6 caracteres.")
            else:
                change_password(st.session_state["username"], nueva)
                st.session_state.debe_cambiar_password = False
                st.rerun()


def _render_sidebar() -> str:
    with st.sidebar:
        rol = st.session_state.get("rol", "cliente")
        rol_label = {"admin": "Administrador", "operaciones": "Operaciones", "cliente": "Cliente"}.get(rol, rol)

        import os as _os, base64 as _b64
        _base = _os.path.dirname(_os.path.abspath(__file__))
        _logo = next((_os.path.join(_base, f) for f in ["logo.png", "logo.jpeg", "logo.png.jpeg", "logo.jpg"] if _os.path.exists(_os.path.join(_base, f))), "")
        if _logo:
            with open(_logo, "rb") as _f:
                _enc = _b64.b64encode(_f.read()).decode()
            _ext = "jpeg" if _logo.endswith((".jpeg", ".jpg", ".png.jpeg")) else "png"
            st.markdown(
                f"<div style='text-align:center;padding:0.8rem 0 0.2rem'>"
                f"<img src='data:image/{_ext};base64,{_enc}' style='width:110px;max-width:80%;image-rendering:auto'/>"
                f"</div>",
                unsafe_allow_html=True,
            )
        st.markdown(f"""
        <div style="text-align:center;padding:0.3rem 0 0.5rem">
            <div style="color:rgba(255,255,255,0.85);font-size:0.95rem;font-weight:600;margin-top:2px">
                {st.session_state.get("user_display", "")}
            </div>
            <div style="color:rgba(255,255,255,0.55);font-size:0.78rem;margin-top:2px">{rol_label}</div>
        </div>
        """, unsafe_allow_html=True)
        st.divider()

        opciones_base = MENU_BY_ROL.get(rol, ["Prefactura Cliente", "Configuración"])
        opciones_display = [MENU_ICONS.get(o, o) for o in opciones_base]
        display_to_base  = dict(zip(opciones_display, opciones_base))

        sel_display = st.radio("Nav", opciones_display, label_visibility="collapsed")
        menu = display_to_base.get(sel_display, opciones_base[0])

        st.divider()
        if st.button("Cerrar sesión", use_container_width=True):
            for k in list(st.session_state.keys()): del st.session_state[k]
            st.rerun()
    return menu


def _render_filters():
    _ldr = st.empty()
    _ldr.markdown(
        "<div style='text-align:center;padding:1rem;color:#6B21A8;font-size:1rem;'>"
        "🏍️ &nbsp;Cargando empresas...</div>",
        unsafe_allow_html=True,
    )
    df_companies = load_companies()
    _ldr.empty()
    if df_companies.empty:
        st.warning("⚠️ No se pudieron cargar las empresas. Verifica la conexión a ClickHouse.")
        return None, None, None, None, None, False

    st.markdown('<div class="filter-header">🔍 Filtros de consulta</div>', unsafe_allow_html=True)

    modo = st.radio(
        "Buscar por",
        ["🏢  Por empresa", "🔢  Por NIT"],
        horizontal=True,
        label_visibility="collapsed",
    )

    col1, col2, col3, col4 = st.columns([3, 2, 2, 1])

    empresa_display = ""
    search_mode     = ""
    search_value    = ""

    if "empresa" in modo.lower():
        company_map = dict(zip(df_companies["Nombre_Compania"], df_companies["Company_ID"]))
        with col1:
            empresas_sel = st.multiselect(
                "Empresa(s)", options=list(company_map.keys()),
                help="Selecciona una o más empresas para combinar en una sola prefactura",
            )
        search_mode  = "company_id"
        search_value = [company_map[e] for e in empresas_sel]
        if len(empresas_sel) == 1:
            empresa_display = empresas_sel[0]
        elif len(empresas_sel) > 1:
            empresa_display = f"Multiples_{len(empresas_sel)}_Empresas"
        else:
            empresa_display = ""
    else:
        with col1:
            nit = st.text_input("NIT", placeholder="Ej: 900123456-1", help="Ingresa el NIT exacto de la empresa")
        search_mode   = "nit"
        search_value  = nit.strip()
        empresa_display = nit.strip()

    with col2:
        fi = st.date_input("Fecha inicio", value=date.today().replace(day=1))
    with col3:
        ff = st.date_input("Fecha fin",    value=date.today())
    with col4:
        st.write(""); st.write("")
        consultar = st.button("🔍 Consultar", use_container_width=True, type="primary")

    # Nombre personalizado — resetear si cambia la empresa seleccionada
    if empresa_display:
        if st.session_state.get("_last_empresa_sel") != empresa_display:
            st.session_state["_nombre_custom"] = empresa_display
            st.session_state["_last_empresa_sel"] = empresa_display
        nombre_custom = st.text_input(
            "📝 Nombre del archivo y Excel",
            key="_nombre_custom",
            placeholder="Ej: Cruz Verde",
            help="Este nombre se usará en el archivo exportado y como título dentro del Excel",
        )
        if nombre_custom.strip():
            empresa_display = nombre_custom.strip()

    return empresa_display, search_mode, search_value, fi, ff, consultar


def _render_metrics(df: pd.DataFrame, fi: date, ff: date) -> None:
    svc_cols = [c for c in ["Booking_ID", "Fecha_VERDADERA", "Nombre_Compania", "vt_name_es", "GMV"] if c in df.columns]
    df_svc   = df[svc_cols].drop_duplicates() if svc_cols else pd.DataFrame()
    gmv      = float(df_svc["GMV"].sum()) if "GMV" in df_svc.columns else 0.0

    c1, c2, c3, c4 = st.columns(4)
    cards = [
        (c1, "🚚", f"{len(df_svc):,}", "Servicios únicos"),
        (c2, "📦", f"{len(df):,}",     "Total paquetes"),
        (c3, "💰", _fmt_cop(gmv),       "GMV total"),
        (c4, "📅", f"{fi.strftime('%d/%m')} — {ff.strftime('%d/%m/%Y')}", "Periodo"),
    ]
    for col, icon, val, label in cards:
        with col:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-icon">{icon}</div>
                <div class="metric-value">{val}</div>
                <div class="metric-label">{label}</div>
            </div>
            """, unsafe_allow_html=True)
    st.write("")


def _render_preview(df: pd.DataFrame, modulo_key: str, empresa: str, fi: date, ff: date) -> None:
    svc_cols = [c for c in _SERVICIOS_MAP.values() if c in df.columns]
    df_svc   = df[svc_cols].drop_duplicates()
    gmv      = float(df_svc["GMV"].sum()) if "GMV" in df_svc.columns else 0.0
    periodo  = f"{MESES_ES.get(fi.month, '')} {fi.year}"
    gmv_str  = _fmt_cop(gmv)

    with st.expander("📋 Vista previa — Resumen", expanded=False):
        st.markdown(f"""
        <div class="preview-resumen">
            <div class="preview-resumen-header">
                <span class="preview-company">{empresa}</span>
                <span class="preview-period">📅 {periodo}</span>
            </div>
            <table>
                <tr><td>Empresa</td><td class="preview-amount">{empresa}</td></tr>
                <tr><td>Periodo</td><td class="preview-amount">{periodo}</td></tr>
                <tr><td>Pago Servicios</td><td class="preview-amount">{gmv_str}</td></tr>
                <tr class="preview-total">
                    <td><strong>Valor total de la factura:</strong></td>
                    <td class="preview-amount"><strong>{gmv_str}</strong></td>
                </tr>
            </table>
            <div class="preview-footer">
                El plazo para dar el OK o para pedir ajustes es de 3 días
                y de esta manera evitar demoras en el proceso de facturación.
            </div>
        </div>
        """, unsafe_allow_html=True)

    with st.expander(f"📊 Vista previa — Informe Servicios ({len(df_svc):,} filas)", expanded=False):
        if not df_svc.empty:
            q_svc = st.text_input("🔍 Buscar booking", key=f"q_svc_{modulo_key}", placeholder="ID servicio...")
            df_show = df_svc.rename(columns={v: k for k, v in _SERVICIOS_MAP.items()})
            if q_svc:
                mask = df_show.astype(str).apply(lambda c: c.str.contains(q_svc, case=False, na=False)).any(axis=1)
                df_show = df_show[mask]
            st.dataframe(df_show, use_container_width=True, height=300, hide_index=True)

    pkg_cols = [c for c in _PAQUETES_MAP.values() if c in df.columns]
    df_pkg   = df[pkg_cols]
    with st.expander(f"📦 Vista previa — Informe Paquetes ({len(df_pkg):,} filas)", expanded=False):
        q_pkg = st.text_input("🔍 Buscar booking o paquete", key=f"q_pkg_{modulo_key}", placeholder="ID servicio, # paquete...")
        df_show = df_pkg.rename(columns={v: k for k, v in _PAQUETES_MAP.items()})
        if q_pkg:
            mask = df_show.astype(str).apply(lambda c: c.str.contains(q_pkg, case=False, na=False)).any(axis=1)
            df_show = df_show[mask]
        st.dataframe(df_show, use_container_width=True, height=300, hide_index=True)

    if modulo_key == "interna":
        pil_cols = [c for c in _PILOTOS_MAP.values() if c in df.columns]
        df_pil   = df[pil_cols]
        with st.expander(f"🏍 Vista previa — Data Pilotos ({len(df_pil):,} filas)", expanded=False):
            q_pil = st.text_input("🔍 Buscar booking o piloto", key=f"q_pil_{modulo_key}", placeholder="ID servicio, nombre piloto...")
            df_show = df_pil.rename(columns={v: k for k, v in _PILOTOS_MAP.items()})
            if q_pil:
                mask = df_show.astype(str).apply(lambda c: c.str.contains(q_pil, case=False, na=False)).any(axis=1)
                df_show = df_show[mask]
            st.dataframe(df_show, use_container_width=True, height=300, hide_index=True)


_MODULE_META = {
    "prefactura_cliente": ("📊", "Prefactura Cliente", "Genera prefactura lista para enviar al cliente"),
    "prefactura_interna": ("📋", "Prefactura Interna", "Informe interno con datos de pilotos y servicios"),
    "data":               ("🗃️", "Data",               "Exporta la data completa del período seleccionado"),
}

def _page_module(title: str, generator_fn, file_prefix: str, modulo_key: str) -> None:
    meta = _MODULE_META.get(file_prefix, ("📄", title, ""))
    st.markdown(
        f"""<div class="module-banner">
            <div class="module-banner-icon">{meta[0]}</div>
            <div class="module-banner-text">
                <h2>{meta[1]}</h2>
                <p>{meta[2]}</p>
            </div>
        </div>""",
        unsafe_allow_html=True,
    )

    result = _render_filters()
    empresa_display, search_mode, search_value, fi, ff, consultar = result
    if empresa_display is None: return

    if consultar:
        for k in [f"{file_prefix}_excel", f"{file_prefix}_filename",
                  f"{file_prefix}_count",  f"{file_prefix}_df_key",
                  f"{file_prefix}_empresa", f"{file_prefix}_fi", f"{file_prefix}_ff"]:
            st.session_state.pop(k, None)

        if not search_value:
            st.warning("⚠️ Ingresa un valor para buscar.")
            return

        _loader = _show_moto_loader(f"Consultando datos de {empresa_display}...")
        df, error = query_data(search_mode, search_value, fi, ff)
        _loader.empty()

        if error:
            with st.expander("❌ Error al consultar — ver detalle"):
                st.code(error)
            return
        if df is None or df.empty:
            st.warning("⚠️ No se encontraron registros para los filtros seleccionados.")
            return

        _loader2 = _show_moto_loader("Generando archivo Excel...")
        excel_bytes = generator_fn(df, empresa_display, fi, ff)
        _loader2.empty()

        empresa_safe = "".join(c if c.isalnum() else "_" for c in empresa_display).strip("_")
        st.session_state[f"{file_prefix}_excel"]    = excel_bytes
        st.session_state[f"{file_prefix}_filename"] = f"{file_prefix}_{empresa_safe}_{fi}_{ff}.xlsx"
        st.session_state[f"{file_prefix}_count"]    = len(df)
        st.session_state[f"{file_prefix}_df"]       = df
        st.session_state[f"{file_prefix}_empresa"]  = empresa_display
        st.session_state[f"{file_prefix}_fi"]       = fi
        st.session_state[f"{file_prefix}_ff"]       = ff

    if f"{file_prefix}_excel" in st.session_state:
        df_cached = st.session_state.get(f"{file_prefix}_df", pd.DataFrame())
        emp       = st.session_state.get(f"{file_prefix}_empresa", "")
        fi_c      = st.session_state.get(f"{file_prefix}_fi", date.today())
        ff_c      = st.session_state.get(f"{file_prefix}_ff", date.today())
        count     = st.session_state[f"{file_prefix}_count"]
        fname     = st.session_state[f"{file_prefix}_filename"]

        st.divider()

        # KPIs
        _render_metrics(df_cached, fi_c, ff_c)

        # Descarga
        emp_safe = "".join(c if c.isalnum() or c in " _-" else "_" for c in emp).strip() or "export"
        fname_final = f"{file_prefix}_{emp_safe}_{fi_c}_{ff_c}.xlsx"
        col_dl, col_info = st.columns([2, 3])
        with col_dl:
            st.download_button(
                label=f"📥  Descargar Excel — {emp}",
                data=st.session_state[f"{file_prefix}_excel"],
                file_name=fname_final,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with col_info:
            size_kb = len(st.session_state[f"{file_prefix}_excel"]) // 1024
            st.markdown(
                f"<div style='padding:0.6rem;color:#6B7280;font-size:0.88rem;'>"
                f"✅ <strong>{count:,}</strong> registros encontrados &nbsp;·&nbsp; "
                f"Tamaño aprox: <strong>{size_kb} KB</strong></div>",
                unsafe_allow_html=True,
            )

        st.divider()

        # Vistas previas
        if not df_cached.empty:
            st.markdown("#### 👁️ Vista previa")
            _render_preview(df_cached, modulo_key, emp, fi_c, ff_c)


def _page_data() -> None:
    """Página Data: descarga toda la data por rango de fechas. Solo financiero/admin."""
    st.markdown(
        """<div class="module-banner">
            <div class="module-banner-icon">🗃️</div>
            <div class="module-banner-text">
                <h2>Data</h2>
                <p>Exporta la data completa del período seleccionado</p>
            </div>
        </div>""",
        unsafe_allow_html=True,
    )

    col_fi, col_ff, col_btn = st.columns([2, 2, 1])
    with col_fi:
        fi = st.date_input("Fecha inicio", value=date.today().replace(day=1), key="data_fi")
    with col_ff:
        ff = st.date_input("Fecha fin", value=date.today(), key="data_ff")
    with col_btn:
        st.markdown("<div style='margin-top:1.8rem'></div>", unsafe_allow_html=True)
        consultar = st.button("🔍 Consultar", type="primary", use_container_width=True)

    if consultar:
        if fi > ff:
            st.warning("⚠️ La fecha inicio no puede ser mayor a la fecha fin.")
            return
        st.session_state.pop("data_all_excel", None)
        _loader = _show_moto_loader(f"Consultando datos {fi} → {ff}...")
        df, error = query_data_all(fi, ff)
        _loader.empty()
        if error:
            with st.expander("❌ Error al consultar"):
                st.code(error)
            return
        if df is None or df.empty:
            st.warning("⚠️ No se encontraron registros para ese período.")
            return
        _loader2 = _show_moto_loader("Generando Excel...")
        excel_bytes = gen_data_excel(df, f"Todo_{fi}_{ff}", fi, ff)
        _loader2.empty()
        st.session_state["data_all_excel"]    = excel_bytes
        st.session_state["data_all_count"]    = len(df)
        st.session_state["data_all_fi"]       = fi
        st.session_state["data_all_ff"]       = ff

    if "data_all_excel" in st.session_state:
        fi_c  = st.session_state["data_all_fi"]
        ff_c  = st.session_state["data_all_ff"]
        count = st.session_state["data_all_count"]
        st.divider()
        col_dl, col_info = st.columns([2, 3])
        with col_dl:
            st.download_button(
                label=f"📥 Descargar Excel — Todo el período",
                data=st.session_state["data_all_excel"],
                file_name=f"Data_{fi_c}_{ff_c}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )
        with col_info:
            st.success(f"✅ {count:,} registros encontrados · {fi_c} → {ff_c}")


def _page_error_tracker() -> None:
    st.markdown(
        """<div class="module-banner">
            <div class="module-banner-icon">🚨</div>
            <div class="module-banner-text">
                <h2>Pibox Error Tracker</h2>
                <p>Bookings con alertas de control — últimos 3 meses</p>
            </div>
        </div>""",
        unsafe_allow_html=True,
    )

    with st.spinner("Cargando datos de errores..."):
        df_raw = query_error_tracker()

    if df_raw.empty:
        st.warning("⚠️ No se encontraron registros con errores en los últimos 3 meses.")
        return

    # ── Refresh info ────────────────────────────────────────────
    if "last_refresh" in df_raw.columns:
        lr = df_raw["last_refresh"].dropna()
        if not lr.empty:
            st.caption(f"🔄 Última actualización DAG: {lr.iloc[-1]}")

    # ── Calcular mes anterior por defecto ───────────────────────
    from datetime import timedelta
    _hoy = date.today()
    _mes_anterior = (_hoy.replace(day=1) - timedelta(days=1)).strftime("%Y-%m")
    _mes_actual   = _hoy.strftime("%Y-%m")
    _default_key  = _mes_anterior if _hoy.day <= 10 else _mes_actual
    meses_disp = sorted(df_raw["Fecha_VERDADERA"].dropna().astype(str).str[:7].unique().tolist(), reverse=True)
    _default_mes = [_default_key] if _default_key in meses_disp else []

    # ── Filtros ─────────────────────────────────────────────────
    col_f1, col_f2, col_f3, col_f4 = st.columns([2, 2, 2, 2])
    with col_f1:
        sel_mes = st.multiselect("📅 Mes", meses_disp, default=_default_mes)
    with col_f2:
        empresas = sorted(df_raw["Nombre_Compania"].dropna().unique().tolist())
        sel_emp = st.multiselect("🏢 Empresa", empresas, placeholder="Todas las empresas")
    with col_f3:
        error_labels = {v[1]: k for k, v in ERROR_CONTROLS.items()}
        sel_err = st.multiselect("⚠️ Tipo de error", list(error_labels.keys()), placeholder="Todos los errores")
    with col_f4:
        ciudades = sorted(df_raw["Ciudad"].dropna().unique().tolist())
        sel_ciudad = st.multiselect("📍 Ciudad", ciudades, placeholder="Todas las ciudades")

    _AFECTAN_CIERRE = {"CONTROL_300", "Control_balanceados_driver", "Control_balanceados_company", "Control_perdidas"}
    _NO_AFECTAN     = {"CONTROL_CERO", "CONTROL_VALOR_ALTO", "CONTROL_VALOR_BAJO"}

    col_t1, col_t2 = st.columns([2, 3])
    with col_t1:
        incluir_cero = st.toggle("Incluir bookings en cero (GMV = 0)", value=False)
    with col_t2:
        filtro_cierre = st.radio(
            "🔒 Impacto en cierre",
            ["Todos", "Solo afectan cierre", "Solo no afectan cierre"],
            horizontal=True,
        )

    df = df_raw.copy()
    if sel_mes:
        df = df[df["Fecha_VERDADERA"].astype(str).str[:7].isin(sel_mes)]
    if not incluir_cero and "CONTROL_CERO" in df.columns:
        df = df[df["CONTROL_CERO"] == "Booking normal"]
    if filtro_cierre == "Solo afectan cierre":
        cols_cierre = [c for c in _AFECTAN_CIERRE if c in df.columns]
        df = df[df[cols_cierre].apply(lambda col: col != "Booking normal").any(axis=1)]
    elif filtro_cierre == "Solo no afectan cierre":
        cols_no = [c for c in _NO_AFECTAN if c in df.columns]
        df = df[df[cols_no].apply(lambda col: col != "Booking normal").any(axis=1)]
    if sel_emp:
        df = df[df["Nombre_Compania"].isin(sel_emp)]
    if sel_ciudad:
        df = df[df["Ciudad"].isin(sel_ciudad)]
    if sel_err:
        cols_sel = [error_labels[l] for l in sel_err]
        mask = df[cols_sel].apply(lambda col: col != "Booking normal").any(axis=1)
        df = df[mask]

    st.divider()

    # ── KPIs por tipo de error ───────────────────────────────────
    st.markdown("#### Resumen de alertas")
    kpi_cols = st.columns(len(ERROR_CONTROLS))
    for i, (ctrl, (val_col, label)) in enumerate(ERROR_CONTROLS.items()):
        count = int((df[ctrl] != "Booking normal").sum()) if ctrl in df.columns else 0
        valor = df[val_col].sum() if val_col in df.columns else 0
        with kpi_cols[i]:
            st.markdown(
                f"""<div style="background:linear-gradient(135deg,#6B21A8,#4C1D95);border-radius:12px;
                padding:1rem;text-align:center;color:white;min-height:100px">
                <div style="font-size:1.6rem;font-weight:700">{count:,}</div>
                <div style="font-size:0.7rem;opacity:0.85;margin-top:0.2rem">{label}</div>
                <div style="font-size:0.75rem;margin-top:0.3rem;opacity:0.7">${valor:,.0f}</div>
                </div>""",
                unsafe_allow_html=True,
            )

    st.markdown("<div style='margin-top:1rem'></div>", unsafe_allow_html=True)

    # ── Gráfico top 20 por empresa ───────────────────────────────
    st.markdown("#### Top 20 empresas con más errores")
    err_counts = {}
    for ctrl, (_, label) in ERROR_CONTROLS.items():
        if ctrl in df.columns:
            grp = df[df[ctrl] != "Booking normal"].groupby("Nombre_Compania").size()
            err_counts[label] = grp
    if err_counts:
        chart_df = pd.DataFrame(err_counts).fillna(0).astype(int)
        chart_df["Total"] = chart_df.sum(axis=1)
        chart_df = chart_df[chart_df["Total"] > 0].sort_values("Total", ascending=False).head(20).drop(columns="Total")
        st.bar_chart(chart_df, use_container_width=True)

    st.divider()

    # ── Tabla detalle ────────────────────────────────────────────
    st.markdown(f"#### Detalle — {len(df):,} bookings con alertas")

    detail_cols = ["Booking_ID", "Fecha_VERDADERA", "Nombre_Compania", "Ciudad",
                   "vt_name_es", "KAM", "GMV"] + list(ERROR_CONTROLS.keys())
    detail_cols = [c for c in detail_cols if c in df.columns]

    q = st.text_input("🔍 Buscar booking, empresa...", placeholder="ID, empresa, ciudad...")
    df_show = df[detail_cols].copy()
    if q:
        mask = df_show.astype(str).apply(lambda c: c.str.contains(q, case=False, na=False)).any(axis=1)
        df_show = df_show[mask]

    # Renombrar columnas de control para mostrar etiquetas legibles
    rename_map = {k: v[1] for k, v in ERROR_CONTROLS.items()}
    df_show = df_show.rename(columns=rename_map)
    st.dataframe(df_show, use_container_width=True, height=450, hide_index=True)

    # ── Descarga ─────────────────────────────────────────────────
    buf = BytesIO()
    df[detail_cols].to_excel(buf, index=False, sheet_name="Error Tracker")
    buf.seek(0)
    st.download_button(
        "📥 Descargar Excel",
        data=buf.getvalue(),
        file_name=f"Error_Tracker_{date.today()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _page_gestion_usuarios() -> None:
    st.title("👥 Gestión de Usuarios")
    usuarios = load_users()

    df_u = pd.DataFrame([{
        "Usuario": u["username"], "Nombre": u["nombre_completo"],
        "Rol": u["rol"],
        "Estado": "✅ Activo" if u.get("activo", True) else "❌ Inactivo",
        "Cambiar PW": "⚠️ Sí" if u.get("debe_cambiar_password") else "—",
    } for u in usuarios])
    st.dataframe(df_u, use_container_width=True, hide_index=True)
    st.divider()

    tab_gest, tab_crear = st.tabs(["✏️ Gestionar usuario existente", "➕ Crear nuevo usuario"])

    with tab_gest:
        sel = st.selectbox("Seleccionar usuario", [u["username"] for u in usuarios])
        user_data = find_user(sel) or {}
        es_yo = sel == st.session_state.get("username")
        col_edit, col_reset = st.columns(2)

        with col_edit:
            st.subheader("Editar datos")
            with st.form("form_editar"):
                nuevo_nombre = st.text_input("Nombre completo", value=user_data.get("nombre_completo", ""))
                nuevo_rol    = st.selectbox("Rol", ROLES, index=ROLES.index(user_data.get("rol", "cliente")))
                activo_actual = user_data.get("activo", True)
                nuevo_activo  = st.checkbox("Activo", value=activo_actual, disabled=es_yo)
                guardar = st.form_submit_button("Guardar cambios", type="primary")
            if guardar:
                if es_yo and not nuevo_activo:
                    st.error("No puedes desactivarte a ti mismo.")
                elif user_data.get("rol") == "admin" and not nuevo_activo and count_active_admins() <= 1:
                    st.error("Debe existir al menos un administrador activo.")
                else:
                    update_user(sel, {"nombre_completo": nuevo_nombre, "rol": nuevo_rol,
                                      "activo": nuevo_activo if not es_yo else activo_actual})
                    st.success("✅ Usuario actualizado."); st.rerun()

        with col_reset:
            st.subheader("Resetear contraseña")
            with st.form("form_reset"):
                temp_pw      = st.text_input("Contraseña temporal", type="password")
                temp_pw_conf = st.text_input("Confirmar contraseña temporal", type="password")
                resetear = st.form_submit_button("Resetear contraseña", type="primary")
            if resetear:
                if temp_pw != temp_pw_conf:  st.error("Las contraseñas no coinciden.")
                elif len(temp_pw) < 6:        st.warning("Mínimo 6 caracteres.")
                else:
                    reset_password_admin(sel, temp_pw)
                    st.success(f"✅ Contraseña de «{sel}» reseteada.")

    with tab_crear:
        st.subheader("Nuevo usuario")
        with st.form("form_crear"):
            new_user   = st.text_input("Username")
            new_nombre = st.text_input("Nombre completo")
            new_rol    = st.selectbox("Rol", ROLES, index=1)
            new_pw     = st.text_input("Contraseña temporal", type="password")
            new_pw2    = st.text_input("Confirmar contraseña", type="password")
            crear = st.form_submit_button("Crear usuario", type="primary")
        if crear:
            if not new_user or not new_nombre or not new_pw: st.error("Todos los campos son obligatorios.")
            elif new_pw != new_pw2: st.error("Las contraseñas no coinciden.")
            elif len(new_pw) < 6:   st.warning("Mínimo 6 caracteres.")
            else:
                ok, msg = create_user(new_user, new_nombre, new_pw, new_rol)
                if ok: st.success(f"✅ {msg} El usuario deberá cambiar su contraseña al ingresar."); st.rerun()
                else:  st.error(msg)


def _page_configuracion() -> None:
    st.title("⚙️ Configuración")
    st.subheader("Cambiar contraseña")
    with st.form("form_cambiar_pw"):
        actual    = st.text_input("Contraseña actual", type="password")
        nueva     = st.text_input("Nueva contraseña", type="password")
        confirmar = st.text_input("Confirmar nueva contraseña", type="password")
        submitted = st.form_submit_button("Actualizar contraseña", type="primary")
    if submitted:
        username = st.session_state.get("username", "")
        if not verify_password(username, actual):   st.error("❌ La contraseña actual es incorrecta.")
        elif nueva != confirmar:                     st.error("❌ Las contraseñas nuevas no coinciden.")
        elif len(nueva) < 6:                         st.warning("Mínimo 6 caracteres.")
        else: change_password(username, nueva);      st.success("✅ Contraseña actualizada exitosamente.")


def _render_footer() -> None:
    st.markdown(
        '<div class="app-footer">Pibox © 2026 &nbsp;·&nbsp; Portal interno de facturación</div>',
        unsafe_allow_html=True,
    )


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main() -> None:
    st.set_page_config(
        page_title="Pibox Facturación", page_icon="📦",
        layout="wide", initial_sidebar_state="expanded",
    )
    st.markdown(APP_CSS, unsafe_allow_html=True)

    if not st.session_state.get("authenticated"):
        _page_login(); return

    if st.session_state.get("debe_cambiar_password"):
        _page_force_change_password(); return

    menu = _render_sidebar()
    rol  = st.session_state.get("rol", "cliente")

    if menu not in MENU_BY_ROL.get(rol, []):
        st.error("No tienes permiso para acceder a esta sección."); return

    if   menu == "Prefactura Cliente":  _page_module("📊 Prefactura Cliente",  gen_prefactura_cliente,  "Prefactura_Cliente", "cliente")
    elif menu == "Prefactura Interna":  _page_module("📋 Prefactura Interna",  gen_prefactura_interna,  "Prefactura_Interna", "interna")
    elif menu == "Data":                _page_data() if rol in ("financiero", "admin") else _page_module("🗃️ Data", gen_data_excel, "Data", "data")
    elif menu == "Error Tracker":       _page_error_tracker()
    elif menu == "Gestión de Usuarios": _page_gestion_usuarios()
    elif menu == "Configuración":       _page_configuracion()

    _render_footer()


if __name__ == "__main__":
    main()
